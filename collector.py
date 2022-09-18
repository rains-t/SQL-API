
def data_collector(path, pages):
    '''this collector is specific to my SQL database, and excel layout. Accepts singular pages
    or tuples of multiple pages ie 'page3' or ('page3','page5,'page8') ''' 
    import datetime
    from openpyxl import load_workbook
    output_list = []
    #dict to change name of towers to necessary name
    #should not be in function honestly

    try:
        wb = load_workbook(path, data_only=True)
    except:
        print(path,'does not exist.')

    adjustment_dict = {'71st': '71', '101st':'101', '51st': 'BAT',
                       'Baptist':'BAP', 'CLEARWELL':'GST'}

    for page in pages:
        
        #ITERATING MULTIPLE SHEETS INSTEAD OF ONE

        ws = wb[page]


        #column letters from excel
        #could use openpyxl.utils.cell.get_column_letter(index_number)
        #cutting cells down due to not needing all information
        columns = ['B','C','D', 'E', 'F', 'G']
        headers =[] #TIME,71st,101st, etc 
        time = []
        #COLLECTING DESIGNATORS
        for letter in columns:
            for cell in ws[letter][6:7]:
                if not cell.value == None:
                    if not str(cell.value) == 'TIME':

                        if str(cell.value) in adjustment_dict.keys():
                            for k, v in adjustment_dict.items():
                                if cell.value == k:
                                    headers.append(v)
                        else:
                            headers.append(cell.value)
                    
        


        #COLLECTING TIMES
        for letter in columns[0]:
            for cell in ws[letter][7:31]:
                if not cell.value == None:
                    output = str(cell.value)
                    time.append(output)

        # time.append('00:00:00') #had issue with date showing up on last time of the night
        # #easier to fix this way than deal with regex for a single time
        #this data is actually unecessary - coincides with 00:00:00 data for next day
        #and is input as two values for 00:00:00 the day it is input.

        #COLLECTING DATE
        for letter in columns[2]:
            for cell in ws[letter][3:4]:
                date = str(cell.value)
                #converting date to useable format for SQL
                date = datetime.datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
                date = datetime.datetime.strftime(date, '%m/%d/%Y')
                

        
        #putting everything together to kick forward to SQL API program
        item_count = 0
        count = -1 #-1 because it needs to start at index 0 to work as an iteration

        for letter in columns[1:]:
            count += 1
            time_subcount = 0  

            for cell in ws[letter][7:31]: #slices the rows to row 7 - 32
                if not cell.value == None:
                    try:
                        cell.value = float(cell.value)
                        data = (date, time[time_subcount], headers[count], cell.value) 
                        #print(item_count, time[time_subcount])
                        item_count += 1
                        time_subcount +=1
                        output_list.append(data)
                        
                        
                        
                    except:
                        print('Error with data at column:', letter,'\n' 'date and time:', date, time[time_subcount],'.')
                        print('Data returned:', cell.value, 'should be a float.')
                        print('Collector aborting.')
                        return

                    

    print(item_count, 'items read')                
    return output_list
                




# ('01_2020.xlsx', (f'{i}' for i in range(1,31)))




















